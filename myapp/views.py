from django.shortcuts import render
import openpyxl
import os
import io
from django.http import HttpResponse
# import xlsxwriter
from xlsxwriter.workbook import Workbook



def download_excel(request):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=result.xlsx"

    wb = openpyxl.load_workbook("result.xlsx")
    return response

def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {"upload_url": ""})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        worksheet = wb.worksheets[0]

        # getting a particular sheet
        # worksheet = sheets[0]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()

        # iterating over the rows and
        # getting value from each cell in row
        first = 0
        for row in worksheet.iter_rows():
            if(first == 0):
                first += 1
                continue
            sum = 0
            cell_count = 1
            row_val = ""
            col_val = ""
            row_data = list()
            for cell in row:
                cell_count += 1
                row_val = int(str(cell.row))
                col_val = int(str(cell.column))
                if(cell.value == None):
                    break
                sum += int(str(cell.value))
            if(cell_count != 1) and sum != 0:
                worksheet.cell(row=row_val, column=col_val).value = sum
            excel_data.append(row_data)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=result.xlsx"
        wb.save(response)
        return response
        # wb.save("result.xlsx")
        # wb.close()
        # return render(request, 'myapp/index.html', {"excel_data":excel_data, "show_download": True})









